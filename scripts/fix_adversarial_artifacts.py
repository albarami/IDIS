#!/usr/bin/env python3
"""Fix adversarial deals 001-008 artifacts.json to match real committed files.

For each deal:
1. Check what files actually exist under artifacts/
2. Compute sha256 and file_size_bytes from disk
3. Update artifacts.json to reference the actual files with correct hashes
4. For deal_008 (version_drift), generate v1 and v2 PDFs if needed

This ensures artifacts.json entries match real committed binary files.
"""

from __future__ import annotations

import hashlib
import io
import json
from pathlib import Path

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font

# PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

TENANT_ID = "00000000-0000-0000-0000-000000000001"
BASE_DATE = "2026-01-05"

ADVERSARIAL_DEALS = [
    ("deal_001_clean", "CloudMetrics", 1),
    ("deal_002_contradiction", "DataSync", 2),
    ("deal_003_unit_mismatch", "MetricFlow", 3),
    ("deal_004_time_window_mismatch", "TimeBase", 4),
    ("deal_005_missing_evidence", "HealthBridge", 5),
    ("deal_006_calc_conflict", "CalcPro", 6),
    ("deal_007_chain_break", "ChainLink", 7),
    ("deal_008_version_drift", "VersionCorp", 8),
]


def generate_pdf(company_name: str, version: str = "v1") -> bytes:
    """Generate a simple PDF for testing."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"{company_name} - Pitch Deck ({version})", styles["Title"]))
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"Version: {version}", styles["Normal"]))
    story.append(Paragraph("Generated for GDBS-FULL dataset testing", styles["Normal"]))

    doc.build(story)
    return buffer.getvalue()


def generate_xlsx(company_name: str) -> bytes:
    """Generate a simple XLSX for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"

    ws["A1"] = "Company"
    ws["B1"] = company_name
    ws["A1"].font = Font(bold=True)

    ws["A2"] = "Generated"
    ws["B2"] = "GDBS-FULL Dataset"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def compute_file_hash(file_path: Path) -> tuple[str, int]:
    """Compute SHA256 hash and file size."""
    content = file_path.read_bytes()
    sha256 = hashlib.sha256(content).hexdigest()
    return sha256, len(content)


def fix_deal_artifacts(deal_dir: Path, company_name: str, deal_num: int) -> None:
    """Fix artifacts for a single deal."""
    artifacts_dir = deal_dir / "artifacts"
    artifacts_json_path = deal_dir / "artifacts.json"

    deal_id = f"00000000-0000-0000-0002-{deal_num:012d}"

    # Ensure artifacts directory exists
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    # Generate/verify PDF
    pdf_path = artifacts_dir / "pitch_deck.pdf"
    if not pdf_path.exists():
        pdf_bytes = generate_pdf(company_name)
        pdf_path.write_bytes(pdf_bytes)
        print(f"  Generated: {pdf_path.name}")

    # Generate/verify XLSX
    xlsx_path = artifacts_dir / "financials.xlsx"
    if not xlsx_path.exists():
        xlsx_bytes = generate_xlsx(company_name)
        xlsx_path.write_bytes(xlsx_bytes)
        print(f"  Generated: {xlsx_path.name}")

    # Compute hashes
    pdf_sha256, pdf_size = compute_file_hash(pdf_path)
    xlsx_sha256, xlsx_size = compute_file_hash(xlsx_path)

    # Build artifacts.json
    artifacts = [
        {
            "artifact_id": f"00000000-0000-0000-0003-{deal_num:06d}000001",
            "tenant_id": TENANT_ID,
            "deal_id": deal_id,
            "artifact_type": "PITCH_DECK",
            "filename": "pitch_deck.pdf",
            "storage_uri": f"file://datasets/gdbs_full/deals/{deal_dir.name}/artifacts/pitch_deck.pdf",
            "connector_type": "Upload",
            "sha256": pdf_sha256,
            "file_size_bytes": pdf_size,
            "version_label": "v1",
            "ingested_at": f"{BASE_DATE}T09:00:00Z",
            "created_at": f"{BASE_DATE}T09:00:00Z",
            "updated_at": f"{BASE_DATE}T09:00:00Z",
        },
        {
            "artifact_id": f"00000000-0000-0000-0003-{deal_num:06d}000002",
            "tenant_id": TENANT_ID,
            "deal_id": deal_id,
            "artifact_type": "FIN_MODEL",
            "filename": "financials.xlsx",
            "storage_uri": f"file://datasets/gdbs_full/deals/{deal_dir.name}/artifacts/financials.xlsx",
            "connector_type": "Upload",
            "sha256": xlsx_sha256,
            "file_size_bytes": xlsx_size,
            "version_label": "v1",
            "ingested_at": f"{BASE_DATE}T09:05:00Z",
            "created_at": f"{BASE_DATE}T09:05:00Z",
            "updated_at": f"{BASE_DATE}T09:05:00Z",
        },
    ]

    # Special handling for deal_008 (version_drift) - needs v1 and v2 PDFs
    if deal_num == 8:
        # Generate v1 and v2 PDFs
        pdf_v1_path = artifacts_dir / "pitch_deck_v1.pdf"
        pdf_v2_path = artifacts_dir / "pitch_deck_v2.pdf"

        if not pdf_v1_path.exists():
            pdf_v1_bytes = generate_pdf(company_name, "v1")
            pdf_v1_path.write_bytes(pdf_v1_bytes)
            print(f"  Generated: {pdf_v1_path.name}")

        if not pdf_v2_path.exists():
            pdf_v2_bytes = generate_pdf(company_name, "v2")
            pdf_v2_path.write_bytes(pdf_v2_bytes)
            print(f"  Generated: {pdf_v2_path.name}")

        pdf_v1_sha256, pdf_v1_size = compute_file_hash(pdf_v1_path)
        pdf_v2_sha256, pdf_v2_size = compute_file_hash(pdf_v2_path)

        # Replace single PDF artifact with v1 and v2
        artifacts = [
            {
                "artifact_id": f"00000000-0000-0000-0003-{deal_num:06d}000001",
                "tenant_id": TENANT_ID,
                "deal_id": deal_id,
                "artifact_type": "PITCH_DECK",
                "filename": "pitch_deck_v1.pdf",
                "storage_uri": f"file://datasets/gdbs_full/deals/{deal_dir.name}/artifacts/pitch_deck_v1.pdf",
                "connector_type": "Upload",
                "sha256": pdf_v1_sha256,
                "file_size_bytes": pdf_v1_size,
                "version_label": "v1",
                "ingested_at": f"{BASE_DATE}T09:00:00Z",
                "created_at": f"{BASE_DATE}T09:00:00Z",
                "updated_at": f"{BASE_DATE}T09:00:00Z",
            },
            {
                "artifact_id": f"00000000-0000-0000-0003-{deal_num:06d}000003",
                "tenant_id": TENANT_ID,
                "deal_id": deal_id,
                "artifact_type": "PITCH_DECK",
                "filename": "pitch_deck_v2.pdf",
                "storage_uri": f"file://datasets/gdbs_full/deals/{deal_dir.name}/artifacts/pitch_deck_v2.pdf",
                "connector_type": "Upload",
                "sha256": pdf_v2_sha256,
                "file_size_bytes": pdf_v2_size,
                "version_label": "v2",
                "ingested_at": f"{BASE_DATE}T10:00:00Z",
                "created_at": f"{BASE_DATE}T10:00:00Z",
                "updated_at": f"{BASE_DATE}T10:00:00Z",
            },
            {
                "artifact_id": f"00000000-0000-0000-0003-{deal_num:06d}000002",
                "tenant_id": TENANT_ID,
                "deal_id": deal_id,
                "artifact_type": "FIN_MODEL",
                "filename": "financials.xlsx",
                "storage_uri": f"file://datasets/gdbs_full/deals/{deal_dir.name}/artifacts/financials.xlsx",
                "connector_type": "Upload",
                "sha256": xlsx_sha256,
                "file_size_bytes": xlsx_size,
                "version_label": "v1",
                "ingested_at": f"{BASE_DATE}T09:05:00Z",
                "created_at": f"{BASE_DATE}T09:05:00Z",
                "updated_at": f"{BASE_DATE}T09:05:00Z",
            },
        ]

    # Write artifacts.json
    artifacts_json_path.write_text(json.dumps({"artifacts": artifacts}, indent=2), encoding="utf-8")
    print(f"  Updated: artifacts.json ({len(artifacts)} artifacts)")


def main() -> None:
    """Fix all adversarial deals artifacts."""
    script_dir = Path(__file__).parent
    repo_root = script_dir.parent
    deals_dir = repo_root / "datasets" / "gdbs_full" / "deals"

    print("Fixing adversarial deals 001-008 artifacts...")
    print("=" * 50)

    for deal_dir_name, company_name, deal_num in ADVERSARIAL_DEALS:
        deal_dir = deals_dir / deal_dir_name
        if not deal_dir.exists():
            print(f"WARN: {deal_dir_name} not found, skipping")
            continue

        print(f"\n{deal_dir_name}:")
        fix_deal_artifacts(deal_dir, company_name, deal_num)

    print("\n" + "=" * 50)
    print("DONE: All adversarial deal artifacts fixed")


if __name__ == "__main__":
    main()
