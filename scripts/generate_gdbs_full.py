#!/usr/bin/env python3
"""
Deterministic GDBS-FULL Dataset Generator.

Generates 100 deals with real PDF/XLSX artifacts for IDIS v6.3.
Uses fixed seed for reproducibility. NO randomness outside seeded RNG.

Usage:
    python scripts/generate_gdbs_full.py
"""

from __future__ import annotations

import hashlib
import io
import json
import random
from pathlib import Path
from typing import Any

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Constants
SEED = 20260109  # Fixed seed for determinism
TENANT_ID = "00000000-0000-0000-0000-000000000001"
ANALYST_1_ID = "00000000-0000-0000-0001-000000000001"
BASE_DATE = "2026-01-05"

# Company name components for deterministic generation
COMPANY_PREFIXES = [
    "Cloud",
    "Data",
    "AI",
    "Quantum",
    "Edge",
    "Cyber",
    "Smart",
    "Digital",
    "Neural",
    "Flex",
    "Swift",
    "Prime",
    "Core",
    "Next",
    "Meta",
    "Alpha",
    "Beta",
    "Gamma",
    "Delta",
    "Omega",
    "Nova",
    "Stellar",
    "Apex",
    "Peak",
    "Summit",
    "Vertex",
    "Zenith",
    "Pulse",
    "Wave",
    "Flow",
    "Stream",
    "Sync",
]
COMPANY_SUFFIXES = [
    "Labs",
    "Systems",
    "Tech",
    "AI",
    "Analytics",
    "Solutions",
    "Dynamics",
    "Insights",
    "Logic",
    "Works",
    "Hub",
    "Space",
    "Stack",
    "Forge",
    "Mind",
    "Sense",
    "Shift",
    "Scale",
    "Grid",
    "Net",
    "Link",
    "Ops",
    "Ware",
    "Base",
]

SECTORS = [
    "SaaS",
    "FinTech",
    "HealthTech",
    "EdTech",
    "DeepTech",
    "CleanTech",
    "Cybersecurity",
    "AI/ML",
]
STAGES = ["SEED", "SERIES_A", "SERIES_B"]


class DeterministicGenerator:
    """Seeded deterministic value generator."""

    def __init__(self, seed: int) -> None:
        self.rng = random.Random(seed)

    def company_name(self, deal_num: int) -> str:
        """Generate deterministic company name."""
        self.rng.seed(SEED + deal_num)
        prefix = self.rng.choice(COMPANY_PREFIXES)
        suffix = self.rng.choice(COMPANY_SUFFIXES)
        return f"{prefix}{suffix}"

    def sector(self, deal_num: int) -> str:
        """Generate deterministic sector."""
        self.rng.seed(SEED + deal_num + 1000)
        return self.rng.choice(SECTORS)

    def stage(self, deal_num: int) -> str:
        """Generate deterministic stage."""
        self.rng.seed(SEED + deal_num + 2000)
        return self.rng.choice(STAGES)

    def arr(self, deal_num: int) -> int:
        """Generate ARR rounded to nearest $1000."""
        self.rng.seed(SEED + deal_num + 3000)
        base = self.rng.randint(1000, 20000)  # $1M to $20M
        return base * 1000

    def gross_margin(self, deal_num: int) -> float:
        """Generate gross margin with 2 decimal places."""
        self.rng.seed(SEED + deal_num + 4000)
        return round(self.rng.uniform(55.0, 85.0), 2)

    def burn(self, deal_num: int) -> int:
        """Generate monthly burn rounded to nearest $100."""
        self.rng.seed(SEED + deal_num + 5000)
        base = self.rng.randint(1000, 8000)  # $100K to $800K
        return base * 100

    def cash(self, deal_num: int) -> int:
        """Generate cash balance."""
        self.rng.seed(SEED + deal_num + 6000)
        return self.rng.randint(5, 30) * 1000000  # $5M to $30M

    def runway(self, deal_num: int, cash: int, burn: int) -> float:
        """Calculate runway with 1 decimal place."""
        if burn == 0:
            return 99.9
        return round(cash / burn, 1)

    def cac(self, deal_num: int) -> int:
        """Generate CAC rounded to nearest $100."""
        self.rng.seed(SEED + deal_num + 7000)
        base = self.rng.randint(10, 80)  # $1000 to $8000
        return base * 100

    def nrr(self, deal_num: int) -> float:
        """Generate NRR with 2 decimal places."""
        self.rng.seed(SEED + deal_num + 8000)
        return round(self.rng.uniform(100.0, 130.0), 2)

    def tam(self, deal_num: int) -> int:
        """Generate TAM rounded to nearest $1M."""
        self.rng.seed(SEED + deal_num + 9000)
        return self.rng.randint(5, 100) * 1000000000  # $5B to $100B


def generate_pdf(
    company_name: str,
    arr: int,
    gm: float,
    burn: int,
    runway: float,
    cac: int,
    nrr: float,
    tam: int,
) -> bytes:
    """Generate a real PDF pitch deck summary."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        spaceAfter=20,
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontSize=12,
        spaceAfter=30,
        alignment=1,
        textColor=colors.grey,
    )

    elements = []

    # Title
    elements.append(Paragraph(f"{company_name}", title_style))
    elements.append(Paragraph("Series A Investment Summary", subtitle_style))
    elements.append(Spacer(1, 20))

    # Key Metrics Table
    metrics_data = [
        ["Metric", "Value", "Period"],
        ["ARR", f"${arr:,}", "FY2025"],
        ["Gross Margin", f"{gm:.2f}%", "FY2025"],
        ["Monthly Burn", f"${burn:,}", "Dec 2025"],
        ["Runway", f"{runway:.1f} months", "Current"],
        ["CAC", f"${cac:,}", "Q4 2025"],
        ["NRR", f"{nrr:.2f}%", "FY2025"],
        ["TAM", f"${tam:,}", "2025 Est."],
    ]

    table = Table(metrics_data, colWidths=[2.5 * inch, 2 * inch, 1.5 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ECF0F1")),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#BDC3C7")),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 11),
                ("TOPPADDING", (0, 1), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 30))

    # Footer
    footer_style = ParagraphStyle(
        "Footer",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey,
        alignment=1,
    )
    elements.append(Paragraph("Confidential - For Investment Committee Review Only", footer_style))
    elements.append(Paragraph(f"Generated: {BASE_DATE}", footer_style))

    doc.build(elements)
    return buffer.getvalue()


def generate_xlsx(
    company_name: str,
    arr: int,
    gm: float,
    burn: int,
    cash: int,
    runway: float,
    cac: int,
    nrr: float,
    tam: int,
    revenue: int | None = None,
    cogs: int | None = None,
) -> bytes:
    """Generate a real XLSX financial model."""
    wb = Workbook()

    # Sheet 1: P&L
    ws_pl = wb.active
    ws_pl.title = "P&L"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Calculate revenue and COGS if not provided
    if revenue is None:
        revenue = arr
    if cogs is None:
        cogs = int(revenue * (1 - gm / 100))

    pl_data = [
        ["P&L Summary", "", "FY2025"],
        ["Revenue", "", f"${revenue:,}"],
        ["COGS", "", f"${cogs:,}"],
        ["Gross Profit", "", f"${revenue - cogs:,}"],
        ["Gross Margin", "", f"{gm:.2f}%"],
        ["Operating Expenses", "", f"${burn * 12:,}"],
        ["EBITDA", "", f"${revenue - cogs - burn * 12:,}"],
    ]

    for row_idx, row_data in enumerate(pl_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_pl.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill

    ws_pl.column_dimensions["A"].width = 25
    ws_pl.column_dimensions["B"].width = 5
    ws_pl.column_dimensions["C"].width = 20

    # Sheet 2: KPI
    ws_kpi = wb.create_sheet("KPI")

    kpi_data = [
        ["KPI Dashboard", "", "Value", "Period"],
        ["ARR", "", f"${arr:,}", "FY2025"],
        ["Gross Margin", "", f"{gm:.2f}%", "FY2025"],
        ["Monthly Burn", "", f"${burn:,}", "Dec 2025"],
        ["Cash Balance", "", f"${cash:,}", "Dec 2025"],
        ["Runway", "", f"{runway:.1f}", "months"],
        ["CAC", "", f"${cac:,}", "Q4 2025"],
        ["NRR", "", f"{nrr:.2f}%", "FY2025"],
        ["TAM", "", f"${tam:,}", "2025"],
    ]

    for row_idx, row_data in enumerate(kpi_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_kpi.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill

    ws_kpi.column_dimensions["A"].width = 20
    ws_kpi.column_dimensions["B"].width = 5
    ws_kpi.column_dimensions["C"].width = 20
    ws_kpi.column_dimensions["D"].width = 15

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def sha256_hex(data: bytes) -> str:
    """Compute SHA256 hash of bytes."""
    return hashlib.sha256(data).hexdigest()


def generate_deal_id(deal_num: int) -> str:
    """Generate deterministic deal UUID."""
    return f"00000000-0000-0000-0002-{deal_num:012d}"


def generate_artifact_id(deal_num: int, artifact_num: int) -> str:
    """Generate deterministic artifact UUID."""
    return f"00000000-0000-0000-0003-{deal_num:06d}{artifact_num:06d}"


def generate_span_id(deal_num: int, span_num: int) -> str:
    """Generate deterministic span UUID."""
    return f"00000000-0000-0000-0004-{deal_num:06d}{span_num:06d}"


def generate_claim_id(deal_num: int, claim_num: int) -> str:
    """Generate deterministic claim UUID."""
    return f"00000000-0000-0000-0005-{deal_num:06d}{claim_num:06d}"


def generate_evidence_id(deal_num: int, evidence_num: int) -> str:
    """Generate deterministic evidence UUID."""
    return f"00000000-0000-0000-0006-{deal_num:06d}{evidence_num:06d}"


def generate_sanad_id(deal_num: int, sanad_num: int) -> str:
    """Generate deterministic sanad UUID."""
    return f"00000000-0000-0000-0007-{deal_num:06d}{sanad_num:06d}"


def generate_calc_id(deal_num: int, calc_num: int) -> str:
    """Generate deterministic calc UUID."""
    return f"00000000-0000-0000-0008-{deal_num:06d}{calc_num:06d}"


def generate_defect_id(deal_num: int, defect_num: int) -> str:
    """Generate deterministic defect UUID."""
    return f"00000000-0000-0000-0009-{deal_num:06d}{defect_num:06d}"


def create_clean_deal(
    deal_num: int,
    gen: DeterministicGenerator,
    output_dir: Path,
) -> dict[str, Any]:
    """Create a clean deal with all components."""
    deal_key = f"deal_{deal_num:03d}"
    deal_dir_name = f"deal_{deal_num:03d}_clean"
    deal_dir = output_dir / "deals" / deal_dir_name
    deal_dir.mkdir(parents=True, exist_ok=True)
    artifacts_dir = deal_dir / "artifacts"
    artifacts_dir.mkdir(exist_ok=True)

    # Generate values
    company_name = gen.company_name(deal_num)
    sector = gen.sector(deal_num)
    stage = gen.stage(deal_num)
    arr = gen.arr(deal_num)
    gm = gen.gross_margin(deal_num)
    burn = gen.burn(deal_num)
    cash = gen.cash(deal_num)
    runway = gen.runway(deal_num, cash, burn)
    cac = gen.cac(deal_num)
    nrr = gen.nrr(deal_num)
    tam = gen.tam(deal_num)

    # Calculate revenue and COGS for calcs
    revenue = arr
    cogs = int(revenue * (1 - gm / 100))

    deal_id = generate_deal_id(deal_num)

    # Generate PDF
    pdf_bytes = generate_pdf(company_name, arr, gm, burn, runway, cac, nrr, tam)
    pdf_path = artifacts_dir / "pitch_deck.pdf"
    pdf_path.write_bytes(pdf_bytes)
    pdf_sha256 = sha256_hex(pdf_bytes)
    pdf_size = len(pdf_bytes)

    # Generate XLSX
    xlsx_bytes = generate_xlsx(
        company_name, arr, gm, burn, cash, runway, cac, nrr, tam, revenue, cogs
    )
    xlsx_path = artifacts_dir / "financials.xlsx"
    xlsx_path.write_bytes(xlsx_bytes)
    xlsx_sha256 = sha256_hex(xlsx_bytes)
    xlsx_size = len(xlsx_bytes)

    # Create deal.json
    deal_json = {
        "deal_id": deal_id,
        "tenant_id": TENANT_ID,
        "external_ref": f"QAV-2026-{deal_num:03d}",
        "company_name": company_name,
        "stage": stage,
        "sector": sector,
        "status": "IN_REVIEW",
        "materiality_threshold": 0.5,
        "scenario": "clean",
        "scenario_description": "Clean baseline deal with all claims verified",
        "created_at": f"{BASE_DATE}T09:00:00Z",
        "updated_at": f"{BASE_DATE}T09:00:00Z",
    }
    (deal_dir / "deal.json").write_text(json.dumps(deal_json, indent=2), encoding="utf-8")

    # Create artifacts.json
    artifacts_json = {
        "artifacts": [
            {
                "artifact_id": generate_artifact_id(deal_num, 1),
                "tenant_id": TENANT_ID,
                "deal_id": deal_id,
                "artifact_type": "PITCH_DECK",
                "filename": "pitch_deck.pdf",
                "storage_uri": f"file://datasets/gdbs_full/deals/{deal_dir_name}/artifacts/pitch_deck.pdf",
                "connector_type": "Upload",
                "sha256": pdf_sha256,
                "file_size_bytes": pdf_size,
                "ingested_at": f"{BASE_DATE}T09:00:00Z",
                "created_at": f"{BASE_DATE}T09:00:00Z",
                "updated_at": f"{BASE_DATE}T09:00:00Z",
            },
            {
                "artifact_id": generate_artifact_id(deal_num, 2),
                "tenant_id": TENANT_ID,
                "deal_id": deal_id,
                "artifact_type": "FIN_MODEL",
                "filename": "financials.xlsx",
                "storage_uri": f"file://datasets/gdbs_full/deals/{deal_dir_name}/artifacts/financials.xlsx",
                "connector_type": "Upload",
                "sha256": xlsx_sha256,
                "file_size_bytes": xlsx_size,
                "ingested_at": f"{BASE_DATE}T09:05:00Z",
                "created_at": f"{BASE_DATE}T09:05:00Z",
                "updated_at": f"{BASE_DATE}T09:05:00Z",
            },
        ]
    }
    (deal_dir / "artifacts.json").write_text(json.dumps(artifacts_json, indent=2), encoding="utf-8")

    # Create spans.json - manual spans referencing real artifact content
    spans_json = {
        "spans": [
            {
                "span_id": generate_span_id(deal_num, 1),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 1),
                "span_type": "PAGE_TEXT",
                "locator": {"page": 1, "bbox": [100, 200, 400, 250]},
                "text_excerpt": f"ARR: ${arr:,}",
                "created_at": f"{BASE_DATE}T09:10:00Z",
                "updated_at": f"{BASE_DATE}T09:10:00Z",
            },
            {
                "span_id": generate_span_id(deal_num, 2),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 1),
                "span_type": "PAGE_TEXT",
                "locator": {"page": 1, "bbox": [100, 260, 400, 310]},
                "text_excerpt": f"Gross Margin: {gm:.2f}%",
                "created_at": f"{BASE_DATE}T09:10:00Z",
                "updated_at": f"{BASE_DATE}T09:10:00Z",
            },
            {
                "span_id": generate_span_id(deal_num, 3),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 1),
                "span_type": "PAGE_TEXT",
                "locator": {"page": 1, "bbox": [100, 320, 400, 370]},
                "text_excerpt": f"Monthly Burn: ${burn:,}",
                "created_at": f"{BASE_DATE}T09:10:00Z",
                "updated_at": f"{BASE_DATE}T09:10:00Z",
            },
            {
                "span_id": generate_span_id(deal_num, 4),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 1),
                "span_type": "PAGE_TEXT",
                "locator": {"page": 1, "bbox": [100, 380, 400, 430]},
                "text_excerpt": f"Runway: {runway:.1f} months",
                "created_at": f"{BASE_DATE}T09:10:00Z",
                "updated_at": f"{BASE_DATE}T09:10:00Z",
            },
            {
                "span_id": generate_span_id(deal_num, 5),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 1),
                "span_type": "PAGE_TEXT",
                "locator": {"page": 1, "bbox": [100, 440, 400, 490]},
                "text_excerpt": f"CAC: ${cac:,}",
                "created_at": f"{BASE_DATE}T09:10:00Z",
                "updated_at": f"{BASE_DATE}T09:10:00Z",
            },
            {
                "span_id": generate_span_id(deal_num, 6),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 1),
                "span_type": "PAGE_TEXT",
                "locator": {"page": 1, "bbox": [100, 500, 400, 550]},
                "text_excerpt": f"NRR: {nrr:.2f}%",
                "created_at": f"{BASE_DATE}T09:10:00Z",
                "updated_at": f"{BASE_DATE}T09:10:00Z",
            },
            {
                "span_id": generate_span_id(deal_num, 7),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 1),
                "span_type": "PAGE_TEXT",
                "locator": {"page": 1, "bbox": [100, 560, 400, 610]},
                "text_excerpt": f"TAM: ${tam:,}",
                "created_at": f"{BASE_DATE}T09:10:00Z",
                "updated_at": f"{BASE_DATE}T09:10:00Z",
            },
            # XLSX spans for corroboration
            {
                "span_id": generate_span_id(deal_num, 8),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 2),
                "span_type": "CELL",
                "locator": {"sheet": "KPI", "cell": "C2"},
                "text_excerpt": f"${arr:,}",
                "created_at": f"{BASE_DATE}T09:15:00Z",
                "updated_at": f"{BASE_DATE}T09:15:00Z",
            },
            {
                "span_id": generate_span_id(deal_num, 9),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 2),
                "span_type": "CELL",
                "locator": {"sheet": "KPI", "cell": "C3"},
                "text_excerpt": f"{gm:.2f}%",
                "created_at": f"{BASE_DATE}T09:15:00Z",
                "updated_at": f"{BASE_DATE}T09:15:00Z",
            },
            {
                "span_id": generate_span_id(deal_num, 10),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 2),
                "span_type": "CELL",
                "locator": {"sheet": "P&L", "cell": "C2"},
                "text_excerpt": f"${revenue:,}",
                "created_at": f"{BASE_DATE}T09:15:00Z",
                "updated_at": f"{BASE_DATE}T09:15:00Z",
            },
            {
                "span_id": generate_span_id(deal_num, 11),
                "tenant_id": TENANT_ID,
                "document_id": generate_artifact_id(deal_num, 2),
                "span_type": "CELL",
                "locator": {"sheet": "P&L", "cell": "C3"},
                "text_excerpt": f"${cogs:,}",
                "created_at": f"{BASE_DATE}T09:15:00Z",
                "updated_at": f"{BASE_DATE}T09:15:00Z",
            },
        ]
    }
    (deal_dir / "spans.json").write_text(json.dumps(spans_json, indent=2), encoding="utf-8")

    # Create claims.json - all 7 claims C1-C7
    claims = [
        {
            "claim_id": generate_claim_id(deal_num, 1),
            "tenant_id": TENANT_ID,
            "deal_id": deal_id,
            "claim_key": "C1",
            "claim_class": "FINANCIAL",
            "claim_text": f"ARR is ${arr:,} as of December 2025",
            "predicate": f"ARR({company_name}, 2025-12) = {arr} USD",
            "value": {
                "value": arr,
                "unit": "USD",
                "currency": "USD",
                "as_of": "2025-12-31",
                "time_window": {
                    "label": "FY2025",
                    "start_date": "2025-01-01",
                    "end_date": "2025-12-31",
                },
            },
            "materiality": "CRITICAL",
            "ic_bound": True,
            "primary_span_id": generate_span_id(deal_num, 1),
            "created_by": ANALYST_1_ID,
            "created_at": f"{BASE_DATE}T10:00:00Z",
            "updated_at": f"{BASE_DATE}T10:00:00Z",
        },
        {
            "claim_id": generate_claim_id(deal_num, 2),
            "tenant_id": TENANT_ID,
            "deal_id": deal_id,
            "claim_key": "C2",
            "claim_class": "FINANCIAL",
            "claim_text": f"Gross Margin is {gm:.2f}%",
            "predicate": f"GrossMargin({company_name}, 2025-12) = {gm}%",
            "value": {
                "value": gm,
                "unit": "percent",
                "currency": None,
                "as_of": "2025-12-31",
                "time_window": {
                    "label": "FY2025",
                    "start_date": "2025-01-01",
                    "end_date": "2025-12-31",
                },
            },
            "materiality": "HIGH",
            "ic_bound": True,
            "primary_span_id": generate_span_id(deal_num, 2),
            "created_by": ANALYST_1_ID,
            "created_at": f"{BASE_DATE}T10:00:00Z",
            "updated_at": f"{BASE_DATE}T10:00:00Z",
        },
        {
            "claim_id": generate_claim_id(deal_num, 3),
            "tenant_id": TENANT_ID,
            "deal_id": deal_id,
            "claim_key": "C3",
            "claim_class": "FINANCIAL",
            "claim_text": f"Monthly Burn Rate is ${burn:,}",
            "predicate": f"MonthlyBurn({company_name}, 2025-12) = {burn} USD",
            "value": {
                "value": burn,
                "unit": "USD/month",
                "currency": "USD",
                "as_of": "2025-12-31",
                "time_window": {
                    "label": "Dec 2025",
                    "start_date": "2025-12-01",
                    "end_date": "2025-12-31",
                },
            },
            "materiality": "HIGH",
            "ic_bound": True,
            "primary_span_id": generate_span_id(deal_num, 3),
            "created_by": ANALYST_1_ID,
            "created_at": f"{BASE_DATE}T10:00:00Z",
            "updated_at": f"{BASE_DATE}T10:00:00Z",
        },
        {
            "claim_id": generate_claim_id(deal_num, 4),
            "tenant_id": TENANT_ID,
            "deal_id": deal_id,
            "claim_key": "C4",
            "claim_class": "FINANCIAL",
            "claim_text": f"Runway is {runway:.1f} months",
            "predicate": f"Runway({company_name}, 2025-12) = {runway} months",
            "value": {
                "value": runway,
                "unit": "months",
                "currency": None,
                "as_of": "2025-12-31",
                "time_window": None,
            },
            "materiality": "CRITICAL",
            "ic_bound": True,
            "primary_span_id": generate_span_id(deal_num, 4),
            "created_by": ANALYST_1_ID,
            "created_at": f"{BASE_DATE}T10:00:00Z",
            "updated_at": f"{BASE_DATE}T10:00:00Z",
        },
        {
            "claim_id": generate_claim_id(deal_num, 5),
            "tenant_id": TENANT_ID,
            "deal_id": deal_id,
            "claim_key": "C5",
            "claim_class": "TRACTION",
            "claim_text": f"CAC is ${cac:,}",
            "predicate": f"CAC({company_name}, 2025-12) = {cac} USD",
            "value": {
                "value": cac,
                "unit": "USD",
                "currency": "USD",
                "as_of": "2025-12-31",
                "time_window": {
                    "label": "Q4 2025",
                    "start_date": "2025-10-01",
                    "end_date": "2025-12-31",
                },
            },
            "materiality": "MEDIUM",
            "ic_bound": True,
            "primary_span_id": generate_span_id(deal_num, 5),
            "created_by": ANALYST_1_ID,
            "created_at": f"{BASE_DATE}T10:00:00Z",
            "updated_at": f"{BASE_DATE}T10:00:00Z",
        },
        {
            "claim_id": generate_claim_id(deal_num, 6),
            "tenant_id": TENANT_ID,
            "deal_id": deal_id,
            "claim_key": "C6",
            "claim_class": "TRACTION",
            "claim_text": f"NRR is {nrr:.2f}%",
            "predicate": f"NRR({company_name}, 2025) = {nrr}%",
            "value": {
                "value": nrr,
                "unit": "percent",
                "currency": None,
                "as_of": "2025-12-31",
                "time_window": {
                    "label": "FY2025",
                    "start_date": "2025-01-01",
                    "end_date": "2025-12-31",
                },
            },
            "materiality": "HIGH",
            "ic_bound": True,
            "primary_span_id": generate_span_id(deal_num, 6),
            "created_by": ANALYST_1_ID,
            "created_at": f"{BASE_DATE}T10:00:00Z",
            "updated_at": f"{BASE_DATE}T10:00:00Z",
        },
        {
            "claim_id": generate_claim_id(deal_num, 7),
            "tenant_id": TENANT_ID,
            "deal_id": deal_id,
            "claim_key": "C7",
            "claim_class": "MARKET_SIZE",
            "claim_text": f"TAM is ${tam:,}",
            "predicate": f"TAM({sector}, 2025) = {tam} USD",
            "value": {
                "value": tam,
                "unit": "USD",
                "currency": "USD",
                "as_of": "2025-12-31",
                "time_window": None,
            },
            "materiality": "MEDIUM",
            "ic_bound": True,
            "primary_span_id": generate_span_id(deal_num, 7),
            "created_by": ANALYST_1_ID,
            "created_at": f"{BASE_DATE}T10:00:00Z",
            "updated_at": f"{BASE_DATE}T10:00:00Z",
        },
    ]
    (deal_dir / "claims.json").write_text(
        json.dumps({"claims": claims}, indent=2), encoding="utf-8"
    )

    # Create evidence.json
    evidence_items = []
    for i in range(1, 8):
        evidence_items.append(
            {
                "evidence_id": generate_evidence_id(deal_num, i),
                "tenant_id": TENANT_ID,
                "deal_id": deal_id,
                "source_span_id": generate_span_id(deal_num, i),
                "source_system": "Deck",
                "upstream_origin_id": f"deck-{company_name.lower().replace(' ', '-')}-c{i}",
                "retrieval_timestamp": f"{BASE_DATE}T09:10:00Z",
                "verification_status": "VERIFIED",
                "source_grade": "C",
                "source_rank_subgrade": "C",
                "rationale": {"reason": "Self-reported metric from pitch deck"},
                "created_at": f"{BASE_DATE}T09:10:00Z",
                "updated_at": f"{BASE_DATE}T09:10:00Z",
            }
        )
        # Corroborating evidence from model (for ARR and GM)
        if i <= 2:
            evidence_items.append(
                {
                    "evidence_id": generate_evidence_id(deal_num, i + 100),
                    "tenant_id": TENANT_ID,
                    "deal_id": deal_id,
                    "source_span_id": generate_span_id(deal_num, i + 7),
                    "source_system": "Model",
                    "upstream_origin_id": f"model-{company_name.lower().replace(' ', '-')}-c{i}",
                    "retrieval_timestamp": f"{BASE_DATE}T09:15:00Z",
                    "verification_status": "VERIFIED",
                    "source_grade": "B",
                    "source_rank_subgrade": "B",
                    "rationale": {"reason": "Verified from financial model"},
                    "created_at": f"{BASE_DATE}T09:15:00Z",
                    "updated_at": f"{BASE_DATE}T09:15:00Z",
                }
            )
    (deal_dir / "evidence.json").write_text(
        json.dumps({"evidence_items": evidence_items}, indent=2), encoding="utf-8"
    )

    # Create sanads.json
    sanads = []
    for i in range(1, 8):
        corroborating = []
        grade = "C"
        corroboration_status = "AHAD_1"
        if i <= 2:  # ARR and GM have corroboration
            corroborating = [generate_evidence_id(deal_num, i + 100)]
            grade = "B"
            corroboration_status = "MUTAWATIR"

        sanads.append(
            {
                "sanad_id": generate_sanad_id(deal_num, i),
                "tenant_id": TENANT_ID,
                "claim_id": generate_claim_id(deal_num, i),
                "deal_id": deal_id,
                "primary_evidence_id": generate_evidence_id(deal_num, i),
                "corroborating_evidence_ids": corroborating,
                "extraction_confidence": 0.95,
                "dhabt_score": 0.90,
                "corroboration_status": corroboration_status,
                "sanad_grade": grade,
                "grade_explanation": [
                    {
                        "step": "base_grade",
                        "value": "C",
                        "reason": "Primary evidence from deck is Grade C",
                    },
                    {
                        "step": "corroboration_upgrade",
                        "value": grade,
                        "reason": f"Corroboration status: {corroboration_status}",
                    },
                ],
                "transmission_chain": [
                    {
                        "node_id": f"00000000-0000-0000-0010-{deal_num:06d}{i:06d}",
                        "node_type": "SOURCE",
                        "input_refs": [],
                        "output_ref": generate_evidence_id(deal_num, i),
                        "description": f"Deck C{i} extraction",
                    }
                ],
                "defects": [],
                "created_at": f"{BASE_DATE}T10:30:00Z",
                "updated_at": f"{BASE_DATE}T10:30:00Z",
            }
        )
    (deal_dir / "sanads.json").write_text(
        json.dumps({"sanads": sanads}, indent=2), encoding="utf-8"
    )

    # Create calcs.json - MANDATORY for all deals
    calc_gm = round((revenue - cogs) / revenue * 100, 2) if revenue > 0 else 0.0
    calc_runway = round(cash / burn, 1) if burn > 0 else 0.0

    calcs = {
        "calc_sanads": [
            {
                "calc_sanad_id": generate_calc_id(deal_num, 1),
                "tenant_id": TENANT_ID,
                "deal_id": deal_id,
                "calc_id": f"calc-gm-{deal_num:03d}",
                "calc_type": "GROSS_MARGIN",
                "input_claim_ids": [generate_claim_id(deal_num, 2)],
                "input_min_sanad_grade": "C",
                "inputs": {
                    "revenue": revenue,
                    "cogs": cogs,
                },
                "formula_hash": "sha256:gm_v1_rev_minus_cogs_div_rev",
                "code_version": "idis-calc-service@1.0.0",
                "output": {
                    "gross_margin_percent": calc_gm,
                },
                "reproducibility_hash": hashlib.sha256(
                    f"gm:{revenue}:{cogs}:{calc_gm}".encode()
                ).hexdigest()[:16],
                "calc_grade": "B",
                "explanation": (
                    f"GM = (Rev - COGS) / Rev = ({revenue} - {cogs}) / {revenue} = {calc_gm}%"
                ),
                "created_at": f"{BASE_DATE}T11:00:00Z",
                "updated_at": f"{BASE_DATE}T11:00:00Z",
            },
            {
                "calc_sanad_id": generate_calc_id(deal_num, 2),
                "tenant_id": TENANT_ID,
                "deal_id": deal_id,
                "calc_id": f"calc-runway-{deal_num:03d}",
                "calc_type": "RUNWAY",
                "input_claim_ids": [generate_claim_id(deal_num, 3), generate_claim_id(deal_num, 4)],
                "input_min_sanad_grade": "C",
                "inputs": {
                    "cash_balance": cash,
                    "monthly_burn": burn,
                },
                "formula_hash": "sha256:runway_v1_cash_div_burn",
                "code_version": "idis-calc-service@1.0.0",
                "output": {
                    "runway_months": calc_runway,
                },
                "reproducibility_hash": hashlib.sha256(
                    f"runway:{cash}:{burn}:{calc_runway}".encode()
                ).hexdigest()[:16],
                "calc_grade": "B",
                "explanation": f"Runway = Cash / Burn = {cash} / {burn} = {calc_runway} months",
                "created_at": f"{BASE_DATE}T11:00:00Z",
                "updated_at": f"{BASE_DATE}T11:00:00Z",
            },
        ]
    }
    (deal_dir / "calcs.json").write_text(json.dumps(calcs, indent=2), encoding="utf-8")

    # Return manifest entry
    return {
        "deal_key": deal_key,
        "scenario": "clean",
        "directory": f"deals/{deal_dir_name}",
        "deal_id": deal_id,
        "description": "Clean baseline deal with all claims verified",
    }


def create_expected_outcome_clean(deal_num: int) -> dict[str, Any]:
    """Create expected outcome for a clean deal."""
    deal_key = f"deal_{deal_num:03d}"
    deal_id = generate_deal_id(deal_num)

    return {
        "deal_id": deal_id,
        "deal_key": deal_key,
        "scenario": "clean",
        "expected_claims": {
            "C1": {
                "claim_grade": "B",
                "claim_verdict": "VERIFIED",
                "claim_action": "NONE",
                "defect_count": 0,
            },
            "C2": {
                "claim_grade": "B",
                "claim_verdict": "VERIFIED",
                "claim_action": "NONE",
                "defect_count": 0,
            },
            "C3": {
                "claim_grade": "C",
                "claim_verdict": "VERIFIED",
                "claim_action": "NONE",
                "defect_count": 0,
            },
            "C4": {
                "claim_grade": "C",
                "claim_verdict": "VERIFIED",
                "claim_action": "NONE",
                "defect_count": 0,
            },
            "C5": {
                "claim_grade": "C",
                "claim_verdict": "VERIFIED",
                "claim_action": "NONE",
                "defect_count": 0,
            },
            "C6": {
                "claim_grade": "C",
                "claim_verdict": "VERIFIED",
                "claim_action": "NONE",
                "defect_count": 0,
            },
            "C7": {
                "claim_grade": "C",
                "claim_verdict": "VERIFIED",
                "claim_action": "NONE",
                "defect_count": 0,
            },
        },
        "expected_defects": [],
        "expected_sanad_coverage": 1.0,
        "expected_grade_d_count": 0,
        "expected_ic_bound_claims": 7,
        "validation_rules": [
            {"rule": "all_claims_have_sanad", "expected": True},
            {"rule": "no_fatal_defects", "expected": True},
            {"rule": "no_major_defects", "expected": True},
            {"rule": "all_material_claims_verified", "expected": True},
        ],
    }


def main() -> None:
    """Generate the complete GDBS-FULL dataset with 100 deals."""
    print("GDBS-FULL Dataset Generator")
    print("=" * 50)
    print(f"Seed: {SEED}")
    print("Total deals: 100 (8 adversarial + 92 clean)")
    print()

    repo_root = Path(__file__).parent.parent
    output_dir = repo_root / "datasets" / "gdbs_full"

    gen = DeterministicGenerator(SEED)

    # Track all deals for manifest
    all_deals: list[dict[str, Any]] = []

    # Keep existing 8 adversarial deals (1-8)
    adversarial_deals = [
        {
            "deal_key": "deal_001",
            "scenario": "clean",
            "directory": "deals/deal_001_clean",
            "deal_id": "00000000-0000-0000-0002-000000000001",
            "description": "Clean baseline deal with all claims verified",
        },
        {
            "deal_key": "deal_002",
            "scenario": "contradiction",
            "directory": "deals/deal_002_contradiction",
            "deal_id": "00000000-0000-0000-0002-000000000002",
            "description": "Deck ARR contradicts Model ARR",
        },
        {
            "deal_key": "deal_003",
            "scenario": "unit_mismatch",
            "directory": "deals/deal_003_unit_mismatch",
            "deal_id": "00000000-0000-0000-0002-000000000003",
            "description": "ARR labeled but value is MRR calculation error",
        },
        {
            "deal_key": "deal_004",
            "scenario": "time_window_mismatch",
            "directory": "deals/deal_004_time_window_mismatch",
            "deal_id": "00000000-0000-0000-0002-000000000004",
            "description": "FY2024 vs LTM time window confusion",
        },
        {
            "deal_key": "deal_005",
            "scenario": "missing_evidence",
            "directory": "deals/deal_005_missing_evidence",
            "deal_id": "00000000-0000-0000-0002-000000000005",
            "description": "Claim with no backing evidence span",
        },
        {
            "deal_key": "deal_006",
            "scenario": "calc_conflict",
            "directory": "deals/deal_006_calc_conflict",
            "deal_id": "00000000-0000-0000-0002-000000000006",
            "description": "Calculated GM differs from stated GM",
        },
        {
            "deal_key": "deal_007",
            "scenario": "chain_break",
            "directory": "deals/deal_007_chain_break",
            "deal_id": "00000000-0000-0000-0002-000000000007",
            "description": "Sanad chain has orphaned transmission node",
        },
        {
            "deal_key": "deal_008",
            "scenario": "version_drift",
            "directory": "deals/deal_008_version_drift",
            "deal_id": "00000000-0000-0000-0002-000000000008",
            "description": "Claim cites old document version, newer exists",
        },
    ]
    all_deals.extend(adversarial_deals)

    # Generate artifacts for existing adversarial deals (1-8)
    print("Generating artifacts for adversarial deals 1-8...")
    for deal_num in range(1, 9):
        deal_entry = adversarial_deals[deal_num - 1]
        deal_dir = output_dir / deal_entry["directory"]
        artifacts_dir = deal_dir / "artifacts"
        artifacts_dir.mkdir(parents=True, exist_ok=True)

        # Generate values for this deal
        company_name = gen.company_name(deal_num)
        arr = gen.arr(deal_num)
        gm = gen.gross_margin(deal_num)
        burn = gen.burn(deal_num)
        cash = gen.cash(deal_num)
        runway = gen.runway(deal_num, cash, burn)
        cac = gen.cac(deal_num)
        nrr = gen.nrr(deal_num)
        tam = gen.tam(deal_num)
        revenue = arr
        cogs = int(revenue * (1 - gm / 100))

        # Generate PDF
        pdf_bytes = generate_pdf(company_name, arr, gm, burn, runway, cac, nrr, tam)
        pdf_path = artifacts_dir / "pitch_deck.pdf"
        pdf_path.write_bytes(pdf_bytes)

        # Generate XLSX
        xlsx_bytes = generate_xlsx(
            company_name, arr, gm, burn, cash, runway, cac, nrr, tam, revenue, cogs
        )
        xlsx_path = artifacts_dir / "financials.xlsx"
        xlsx_path.write_bytes(xlsx_bytes)

        # Update artifacts.json with current sha256 and file_size_bytes
        artifacts_json_path = deal_dir / "artifacts.json"
        if artifacts_json_path.exists():
            artifacts_data = json.loads(artifacts_json_path.read_text(encoding="utf-8"))
            for artifact in artifacts_data.get("artifacts", []):
                fname = artifact.get("filename", "")
                fpath = artifacts_dir / fname
                if fpath.exists():
                    content = fpath.read_bytes()
                    artifact["sha256"] = sha256_hex(content)
                    artifact["file_size_bytes"] = len(content)
            artifacts_json_path.write_text(
                json.dumps(artifacts_data, indent=2) + "\n", encoding="utf-8"
            )

        print(f"  Deal {deal_num:03d}: {company_name}")

    # Generate 92 additional clean deals (9-100)
    print("\nGenerating clean deals 9-100...")
    for deal_num in range(9, 101):
        deal_entry = create_clean_deal(deal_num, gen, output_dir)
        all_deals.append(deal_entry)
        if deal_num % 10 == 0:
            print(f"  Generated deals up to {deal_num}")

    # Generate expected outcomes for new clean deals
    print("\nGenerating expected outcomes for deals 9-100...")
    outcomes_dir = output_dir / "expected_outcomes"
    outcomes_dir.mkdir(exist_ok=True)
    for deal_num in range(9, 101):
        outcome = create_expected_outcome_clean(deal_num)
        outcome_path = outcomes_dir / f"deal_{deal_num:03d}_expected.json"
        outcome_path.write_text(json.dumps(outcome, indent=2), encoding="utf-8")

    # Update manifest
    print("\nUpdating manifest.json for 100 deals...")
    manifest_path = output_dir / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["deals"] = all_deals

    # Update expected_outcomes list
    manifest["expected_outcomes"] = [
        f"expected_outcomes/deal_{i:03d}_expected.json" for i in range(1, 101)
    ]

    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    print("\n" + "=" * 50)
    print("GENERATION COMPLETE")
    print(f"Total deals: {len(all_deals)}")
    print("Adversarial deals: 8 (deals 1-8)")
    print("Clean deals: 92 (deals 9-100)")
    print("=" * 50)


if __name__ == "__main__":
    main()
