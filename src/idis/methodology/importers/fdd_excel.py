"""Synthetic-safe Financial DD Excel methodology importer."""

from __future__ import annotations

import hashlib
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from idis.methodology.ids import generate_methodology_question_id, normalize_text
from idis.methodology.models import (
    AssignedAgent,
    MethodologyQuestion,
    MethodologyRegistry,
    MethodologySourceTrace,
    MethodologyType,
    MethodologyVersion,
    RedFlagRule,
    ReportMapping,
    RequiredCalculation,
    RequiredEvidence,
)

REQUIRED_SHEETS = ("P&L", "Cash Flow", "Liabilities", "Assets")
REQUIRED_COLUMNS = (
    "Term",
    "Nature",
    "Financial statement line",
    "Usual Due Diligence Questions",
)


class FDDImporterError(ValueError):
    """Raised when an FDD workbook cannot be imported safely."""


def import_fdd_workbook(workbook: bytes | Path | str | BinaryIO) -> MethodologyRegistry:
    """Import a sanitized FDD-style workbook into a MethodologyRegistry."""
    workbook_bytes, source_name = _read_workbook(workbook)
    source_hash = hashlib.sha256(workbook_bytes).hexdigest()
    try:
        loaded = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise FDDImporterError("invalid workbook shape") from exc

    questions: list[MethodologyQuestion] = []
    seen_ids: set[str] = set()
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in loaded.sheetnames:
            raise FDDImporterError(f"missing required sheet: {sheet_name}")
        sheet = loaded[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise FDDImporterError(f"invalid workbook shape: empty sheet {sheet_name}")
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        column_indexes = _column_indexes(headers, sheet_name)

        for row_number, row in enumerate(rows[1:], start=2):
            raw_question = _cell(row, column_indexes["Usual Due Diligence Questions"])
            normalized_question = normalize_text(raw_question)
            question_text = raw_question.strip()
            if not normalized_question:
                raise FDDImporterError(f"empty question text at {sheet_name}:{row_number}")

            term = _cell(row, column_indexes["Term"])
            nature = _cell(row, column_indexes["Nature"])
            line_item = _cell(row, column_indexes["Financial statement line"])
            question_id = generate_methodology_question_id(
                methodology_type=MethodologyType.FINANCIAL_DD,
                section=sheet_name,
                sheet_or_section=sheet_name,
                row_number=None,
                line_item=line_item,
                question_text=normalized_question,
            )
            if question_id in seen_ids:
                raise FDDImporterError(f"duplicate methodology_question_id: {question_id}")
            seen_ids.add(question_id)

            questions.append(
                MethodologyQuestion(
                    methodology_id="financial_dd",
                    methodology_version_id=f"financial_dd:{source_hash[:12]}",
                    methodology_question_id=question_id,
                    methodology_type=MethodologyType.FINANCIAL_DD,
                    section=sheet_name,
                    sheet_or_source_section=sheet_name,
                    source_row_number=row_number,
                    term=term,
                    nature=nature,
                    line_item=line_item,
                    question_text=question_text,
                    required_evidence=[
                        RequiredEvidence(
                            evidence_type="financial_support",
                            description=f"Evidence supporting {line_item or sheet_name}",
                        )
                    ],
                    target_document_categories=["financial_statement", "workbook_support"],
                    required_calculations=_required_calculations_for_sheet(sheet_name),
                    assigned_agents=[
                        AssignedAgent(
                            role="financial_agent",
                            responsibility=f"Answer FDD question for {sheet_name}",
                        )
                    ],
                    red_flag_rules=[
                        RedFlagRule(
                            rule_id=f"{sheet_name.lower().replace(' ', '_')}_missing_support",
                            description="Question lacks supporting source evidence",
                            severity="HIGH",
                        )
                    ],
                    report_mapping=ReportMapping(report_section=sheet_name),
                    validation_requirements=["requires_claim_or_evidence"],
                    source_trace=MethodologySourceTrace(
                        source_type="synthetic_workbook",
                        source_name=source_name,
                        source_hash=source_hash,
                        sheet_or_section=sheet_name,
                        row_number=row_number,
                    ),
                )
            )

    version = MethodologyVersion(
        methodology_id="financial_dd",
        methodology_version_id=f"financial_dd:{source_hash[:12]}",
        methodology_type=MethodologyType.FINANCIAL_DD,
        version_label="synthetic_excel",
        source_hash=source_hash,
        source_name=source_name,
        questions=sorted(
            questions,
            key=lambda question: (
                REQUIRED_SHEETS.index(question.source_trace.sheet_or_section),
                question.source_trace.row_number or 0,
                question.methodology_question_id,
            ),
        ),
    )
    return MethodologyRegistry(
        methodology_id="financial_dd",
        methodology_type=MethodologyType.FINANCIAL_DD,
        versions=[version],
    )


def _read_workbook(workbook: bytes | Path | str | BinaryIO) -> tuple[bytes, str]:
    if isinstance(workbook, bytes):
        return workbook, "workbook_bytes.xlsx"
    if isinstance(workbook, (Path, str)):
        path = Path(workbook)
        return path.read_bytes(), path.name
    return workbook.read(), "workbook_stream.xlsx"


def _column_indexes(headers: list[str], sheet_name: str) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for required in REQUIRED_COLUMNS:
        if required not in headers:
            raise FDDImporterError(f"missing required column {required} in {sheet_name}")
        indexes[required] = headers.index(required)
    return indexes


def _cell(row: tuple[object, ...], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _required_calculations_for_sheet(sheet_name: str) -> list[RequiredCalculation]:
    if sheet_name == "P&L":
        return [RequiredCalculation(calc_type="GROSS_MARGIN", required=False)]
    if sheet_name == "Cash Flow":
        return [RequiredCalculation(calc_type="BURN_RATE", required=False)]
    return []
