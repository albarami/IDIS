"""XLSX parser — deterministic cell extraction with sheet/cell locators.

Extracts cell values from XLSX files, producing SpanDraft objects
with stable locators including sheet name, A1 notation, and row/col.

Requirements:
- Deterministic: same bytes in → same ordered spans out
- Fail-closed: malformed XLSX files return structured errors
- Numeric stability: floats formatted via Decimal to avoid binary surprises
- Date handling: ISO-8601 format for datetime values
"""

from __future__ import annotations

import hashlib
import io
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

from idis.parsers.base import (
    ParseError,
    ParseErrorCode,
    ParseLimits,
    ParseResult,
    SpanDraft,
)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException


def _compute_content_hash(text: str) -> str:
    """Compute SHA-256 hash of text content."""
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _format_cell_value(value: Any) -> str | None:
    """Format cell value deterministically.

    Returns:
        String representation of the value, or None if cell is empty.

    Formatting rules:
        - None: returns None (empty cell)
        - str: stripped, empty string returns None
        - int: str(value)
        - float: Decimal(str(value)) to avoid binary float surprises
        - datetime: ISO-8601 format
        - date: ISO-8601 date format
        - time: ISO-8601 time format
        - bool: "TRUE" or "FALSE"
        - other: str(value)
    """
    if value is None:
        return None

    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None

    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        decimal_val = Decimal(str(value))
        normalized = decimal_val.normalize()
        if normalized == normalized.to_integral_value():
            return str(int(normalized))
        return str(normalized)

    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, time):
        return value.isoformat()

    return str(value)


def parse_xlsx(
    data: bytes,
    limits: ParseLimits | None = None,
) -> ParseResult:
    """Parse XLSX bytes and extract cell spans with sheet/cell locators.

    Args:
        data: Raw XLSX file bytes.
        limits: Optional parsing limits (defaults to ParseLimits()).

    Returns:
        ParseResult with success=True and spans if extraction succeeded,
        or success=False with structured errors if parsing failed.

    Behavior:
        - Sheets are processed in workbook order.
        - Only non-empty cells are extracted.
        - Each cell becomes a SpanDraft with locator {sheet, cell, row, col}.
        - Numeric values use Decimal formatting for stability.
        - Dates use ISO-8601 format.
        - Malformed XLSX files fail with INVALID_XLSX error.
    """
    if limits is None:
        limits = ParseLimits()

    if len(data) > limits.max_bytes:
        return ParseResult(
            doc_type="XLSX",
            success=False,
            errors=[
                ParseError(
                    code=ParseErrorCode.MAX_SIZE_EXCEEDED,
                    message=f"File size {len(data)} bytes exceeds limit {limits.max_bytes}",
                    details={"size": len(data), "limit": limits.max_bytes},
                )
            ],
        )

    try:
        workbook = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=True,
        )
    except InvalidFileException as e:
        return ParseResult(
            doc_type="XLSX",
            success=False,
            errors=[
                ParseError(
                    code=ParseErrorCode.INVALID_XLSX,
                    message="Invalid XLSX file format",
                    details={"error": str(e)},
                )
            ],
        )
    except Exception as e:
        return ParseResult(
            doc_type="XLSX",
            success=False,
            errors=[
                ParseError(
                    code=ParseErrorCode.CORRUPTED_FILE,
                    message="Failed to read XLSX file",
                    details={"error": str(e), "type": type(e).__name__},
                )
            ],
        )

    sheet_names = workbook.sheetnames
    if len(sheet_names) > limits.max_sheets:
        workbook.close()
        return ParseResult(
            doc_type="XLSX",
            success=False,
            errors=[
                ParseError(
                    code=ParseErrorCode.MAX_SHEETS_EXCEEDED,
                    message=(
                        f"Workbook has {len(sheet_names)} sheets, exceeds limit {limits.max_sheets}"
                    ),
                    details={"sheets": len(sheet_names), "limit": limits.max_sheets},
                )
            ],
        )

    spans: list[SpanDraft] = []
    warnings: list[str] = []
    total_cells = 0

    try:
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            sheet_cells = 0

            for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    cell_value = _format_cell_value(cell.value)
                    if cell_value is None:
                        continue

                    sheet_cells += 1
                    total_cells += 1

                    if sheet_cells > limits.max_cells_per_sheet:
                        warnings.append(
                            f"Sheet '{sheet_name}': exceeded "
                            f"{limits.max_cells_per_sheet} cells, truncated"
                        )
                        break

                    if total_cells > limits.max_total_cells:
                        workbook.close()
                        return ParseResult(
                            doc_type="XLSX",
                            success=False,
                            errors=[
                                ParseError(
                                    code=ParseErrorCode.MAX_CELLS_EXCEEDED,
                                    message=(
                                        f"Total cells {total_cells} exceeds limit "
                                        f"{limits.max_total_cells}"
                                    ),
                                    details={
                                        "cells": total_cells,
                                        "limit": limits.max_total_cells,
                                    },
                                )
                            ],
                        )

                    col_letter = get_column_letter(col_idx)
                    cell_ref = f"{col_letter}{row_idx}"

                    spans.append(
                        SpanDraft(
                            span_type="CELL",
                            locator={
                                "sheet": sheet_name,
                                "cell": cell_ref,
                                "row": row_idx,
                                "col": col_idx,
                            },
                            text_excerpt=cell_value,
                            content_hash=_compute_content_hash(cell_value),
                        )
                    )

                if sheet_cells > limits.max_cells_per_sheet:
                    break
    finally:
        workbook.close()

    return ParseResult(
        doc_type="XLSX",
        success=True,
        spans=spans,
        metadata={
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "span_count": len(spans),
            "total_cells": total_cells,
        },
        warnings=warnings,
    )
