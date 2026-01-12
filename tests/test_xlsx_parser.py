"""Tests for XLSX parser — Phase 1.2.

Tests cover:
- Successful parsing with cell extraction
- Locator correctness (sheet + cell A1 notation + row/col)
- Deterministic numeric formatting
- Date handling (ISO-8601)
- Fail-closed behavior for corrupted files
- Size limit enforcement
"""

from __future__ import annotations

import io
from datetime import date, datetime

import pytest

from idis.parsers.base import ParseErrorCode, ParseLimits
from idis.parsers.xlsx import parse_xlsx

try:
    from openpyxl import Workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def create_test_xlsx(
    sheets: dict[str, list[list[object]]],
) -> bytes:
    """Create an XLSX workbook with given sheet data.

    Args:
        sheets: Dict mapping sheet names to 2D list of cell values.
                Each inner list is a row.

    Returns:
        XLSX file as bytes.
    """
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl not installed")

    wb = Workbook()
    first_sheet = True

    for sheet_name, rows in sheets.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        for row_idx, row_data in enumerate(rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestXLSXParserSuccess:
    """Test successful XLSX parsing scenarios."""

    def test_parse_simple_xlsx(self) -> None:
        """Parse a simple XLSX and verify spans are extracted."""
        sheets = {
            "Sheet1": [
                ["Header A", "Header B"],
                ["Value 1", "Value 2"],
            ]
        }
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        assert result.doc_type == "XLSX"
        assert len(result.errors) == 0
        assert len(result.spans) == 4  # 4 cells with data

    def test_locator_contains_required_fields(self) -> None:
        """Verify locators contain sheet, cell (A1), row, and col."""
        sheets = {"TestSheet": [["Value"]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        for span in result.spans:
            locator = span.locator
            assert "sheet" in locator, "Locator must contain 'sheet'"
            assert "cell" in locator, "Locator must contain 'cell'"
            assert "row" in locator, "Locator must contain 'row'"
            assert "col" in locator, "Locator must contain 'col'"

    def test_locator_cell_is_a1_notation(self) -> None:
        """Verify cell reference uses A1 notation."""
        sheets = {"Sheet1": [["A1Val", "B1Val"], ["A2Val", "B2Val"]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        cell_refs = [span.locator["cell"] for span in result.spans]
        assert "A1" in cell_refs
        assert "B1" in cell_refs
        assert "A2" in cell_refs
        assert "B2" in cell_refs

    def test_locator_row_col_integers(self) -> None:
        """Verify row and col are 1-indexed integers."""
        sheets = {"Sheet1": [["Value"]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        span = result.spans[0]
        assert isinstance(span.locator["row"], int)
        assert isinstance(span.locator["col"], int)
        assert span.locator["row"] >= 1
        assert span.locator["col"] >= 1

    def test_span_type_is_cell(self) -> None:
        """Verify all spans have span_type CELL."""
        sheets = {"Sheet1": [["Value"]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        for span in result.spans:
            assert span.span_type == "CELL"

    def test_spans_have_content_hash(self) -> None:
        """Verify all spans have content_hash populated."""
        sheets = {"Sheet1": [["Value"]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        for span in result.spans:
            assert span.content_hash is not None
            assert len(span.content_hash) == 64  # SHA-256 hex


class TestXLSXParserSpecialSheetNames:
    """Test handling of special sheet names."""

    def test_sheet_name_with_spaces(self) -> None:
        """Sheet names with spaces are preserved."""
        sheets = {"Profit and Loss Statement": [["Revenue", 1000000]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        sheet_names = {span.locator["sheet"] for span in result.spans}
        assert "Profit and Loss Statement" in sheet_names

    def test_sheet_name_with_special_chars(self) -> None:
        """Sheet names with special characters are preserved."""
        sheets = {"P&L (2024)": [["Data", 100]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        sheet_names = {span.locator["sheet"] for span in result.spans}
        assert "P&L (2024)" in sheet_names


class TestXLSXParserDataTypes:
    """Test handling of different data types."""

    def test_numeric_formatting_integer(self) -> None:
        """Integer values are formatted without decimals."""
        sheets = {"Sheet1": [[42]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        assert result.spans[0].text_excerpt == "42"

    def test_numeric_formatting_float(self) -> None:
        """Float values use Decimal formatting (no binary surprises)."""
        sheets = {"Sheet1": [[3.14159]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        text = result.spans[0].text_excerpt
        assert "3.14159" in text

    def test_numeric_formatting_float_whole(self) -> None:
        """Float values that are whole numbers format as integers."""
        sheets = {"Sheet1": [[100.0]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        assert result.spans[0].text_excerpt == "100"

    def test_date_formatting_iso8601(self) -> None:
        """Date values use ISO-8601 format."""
        test_date = date(2024, 6, 15)
        sheets = {"Sheet1": [[test_date]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        assert "2024-06-15" in result.spans[0].text_excerpt

    def test_datetime_formatting_iso8601(self) -> None:
        """Datetime values use ISO-8601 format."""
        test_dt = datetime(2024, 6, 15, 14, 30, 0)
        sheets = {"Sheet1": [[test_dt]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        text = result.spans[0].text_excerpt
        assert "2024-06-15" in text
        assert "14:30" in text

    def test_string_values_preserved(self) -> None:
        """String values are preserved as-is."""
        sheets = {"Sheet1": [["Hello World"]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        assert result.spans[0].text_excerpt == "Hello World"

    def test_boolean_values(self) -> None:
        """Boolean values are formatted as TRUE/FALSE."""
        sheets = {"Sheet1": [[True, False]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        texts = {span.text_excerpt for span in result.spans}
        assert "TRUE" in texts
        assert "FALSE" in texts

    def test_empty_cells_not_included(self) -> None:
        """Empty cells are not included in spans."""
        sheets = {"Sheet1": [["A", None, "C"], [None, "B", None]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        assert len(result.spans) == 3  # Only A, C, B


class TestXLSXParserMultiSheet:
    """Test multi-sheet workbook handling."""

    def test_multiple_sheets_extracted(self) -> None:
        """All sheets are processed."""
        sheets = {
            "Revenue": [["Q1", 1000]],
            "Expenses": [["Rent", 500]],
            "Summary": [["Net", 500]],
        }
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        assert result.metadata["sheet_count"] == 3
        sheet_names = {span.locator["sheet"] for span in result.spans}
        assert "Revenue" in sheet_names
        assert "Expenses" in sheet_names
        assert "Summary" in sheet_names


class TestXLSXParserDeterminism:
    """Test deterministic behavior of XLSX parser."""

    def test_same_bytes_same_output(self) -> None:
        """Parsing same bytes twice produces identical results."""
        sheets = {
            "Sheet1": [
                ["Revenue", 10000000],
                ["Growth", 1.5],
            ]
        }
        xlsx_bytes = create_test_xlsx(sheets)

        result1 = parse_xlsx(xlsx_bytes)
        result2 = parse_xlsx(xlsx_bytes)

        assert result1.success is True
        assert result2.success is True
        assert len(result1.spans) == len(result2.spans)

        for span1, span2 in zip(result1.spans, result2.spans, strict=True):
            assert span1.locator == span2.locator
            assert span1.text_excerpt == span2.text_excerpt
            assert span1.content_hash == span2.content_hash


class TestXLSXParserFailClosed:
    """Test fail-closed behavior for invalid inputs."""

    def test_corrupted_bytes(self) -> None:
        """Corrupted data returns error, no exception."""
        corrupted = b"PK\x03\x04" + b"\x00\xff" * 100

        result = parse_xlsx(corrupted)

        assert result.success is False
        assert len(result.errors) > 0

    def test_random_bytes(self) -> None:
        """Random bytes return error, no exception."""
        random_data = b"This is not an XLSX file at all"

        result = parse_xlsx(random_data)

        assert result.success is False
        assert len(result.errors) > 0

    def test_truncated_xlsx(self) -> None:
        """Truncated XLSX returns error, no exception."""
        sheets = {"Sheet1": [["Data"]]}
        xlsx_bytes = create_test_xlsx(sheets)
        truncated = xlsx_bytes[: len(xlsx_bytes) // 3]

        result = parse_xlsx(truncated)

        assert result.success is False
        assert len(result.errors) > 0


class TestXLSXParserLimits:
    """Test parsing limit enforcement."""

    def test_max_size_exceeded(self) -> None:
        """File exceeding max_bytes limit returns error."""
        sheets = {"Sheet1": [["Data"]]}
        xlsx_bytes = create_test_xlsx(sheets)
        tiny_limit = ParseLimits(max_bytes=100)

        result = parse_xlsx(xlsx_bytes, limits=tiny_limit)

        assert result.success is False
        error_codes = [e.code for e in result.errors]
        assert ParseErrorCode.MAX_SIZE_EXCEEDED in error_codes

    def test_max_sheets_exceeded(self) -> None:
        """Workbook exceeding max_sheets limit returns error."""
        sheets = {f"Sheet{i}": [[i]] for i in range(10)}
        xlsx_bytes = create_test_xlsx(sheets)
        small_limit = ParseLimits(max_sheets=5)

        result = parse_xlsx(xlsx_bytes, limits=small_limit)

        assert result.success is False
        error_codes = [e.code for e in result.errors]
        assert ParseErrorCode.MAX_SHEETS_EXCEEDED in error_codes


class TestXLSXParserMetadata:
    """Test metadata extraction."""

    def test_metadata_contains_sheet_count(self) -> None:
        """Metadata includes sheet_count."""
        sheets = {"Sheet1": [[1]], "Sheet2": [[2]], "Sheet3": [[3]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        assert "sheet_count" in result.metadata
        assert result.metadata["sheet_count"] == 3

    def test_metadata_contains_sheet_names(self) -> None:
        """Metadata includes sheet_names list."""
        sheets = {"Alpha": [[1]], "Beta": [[2]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        assert "sheet_names" in result.metadata
        assert "Alpha" in result.metadata["sheet_names"]
        assert "Beta" in result.metadata["sheet_names"]

    def test_metadata_contains_span_count(self) -> None:
        """Metadata includes span_count."""
        sheets = {"Sheet1": [[1, 2], [3, 4]]}
        xlsx_bytes = create_test_xlsx(sheets)

        result = parse_xlsx(xlsx_bytes)

        assert result.success is True
        assert "span_count" in result.metadata
        assert result.metadata["span_count"] == 4
